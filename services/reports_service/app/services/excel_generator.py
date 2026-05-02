import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from ..schemas import MonthlyRevenueReport


class ExcelGenerator:
    @staticmethod
    def generate_revenue_report(report: MonthlyRevenueReport, hotel_name: str = "Hotel") -> bytes:
        """Genera un archivo Excel del reporte de ingresos mensuales."""
        buffer = io.BytesIO()
        wb = Workbook()

        # Hoja 1: Resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"

        # Título
        ws_summary["A1"] = f"Reporte de Ingresos Mensuales - {hotel_name}"
        ws_summary["A1"].font = Font(size=14, bold=True)
        ws_summary["A2"] = f"Periodo: {report.summary.month}/{report.summary.year}"
        ws_summary["A2"].font = Font(size=11)

        # Resumen financiero
        ws_summary["A4"] = "Concepto"
        ws_summary["B4"] = "Monto"
        ws_summary["A4"].font = Font(bold=True)
        ws_summary["B4"].font = Font(bold=True)
        ws_summary["A4"].fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        ws_summary["B4"].fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        ws_summary["A5"] = "Ingresos Brutos"
        ws_summary["B5"] = float(report.summary.gross_revenue)
        ws_summary["B5"].number_format = f'"${report.summary.currency}" #,##0.00'

        ws_summary["A6"] = "Cancelaciones"
        ws_summary["B6"] = -float(report.summary.cancellations_amount)
        ws_summary["B6"].number_format = f'"${report.summary.currency}" #,##0.00'

        ws_summary["A7"] = "Reembolsos"
        ws_summary["B7"] = -float(report.summary.refunds_amount)
        ws_summary["B7"].number_format = f'"${report.summary.currency}" #,##0.00'

        ws_summary["A8"] = "Ingreso Neto"
        ws_summary["B8"] = float(report.summary.net_revenue)
        ws_summary["B8"].number_format = f'"${report.summary.currency}" #,##0.00'
        ws_summary["B8"].font = Font(bold=True)
        ws_summary["B8"].fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )

        # Estadísticas
        ws_summary["A10"] = "Estadísticas"
        ws_summary["A10"].font = Font(bold=True)
        ws_summary["A11"] = "Total de Reservas"
        ws_summary["B11"] = report.summary.total_bookings
        ws_summary["A12"] = "Confirmadas"
        ws_summary["B12"] = report.summary.confirmed_bookings
        ws_summary["A13"] = "Canceladas"
        ws_summary["B13"] = report.summary.cancelled_bookings
        ws_summary["A14"] = "Pendientes"
        ws_summary["B14"] = report.summary.pending_bookings

        # Ajustar anchos de columna
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20

        # Hoja 2: Transacciones
        if report.transactions:
            ws_trans = wb.create_sheet("Transacciones")

            # Crear DataFrame
            trans_data = []
            for tx in report.transactions:
                trans_data.append(
                    {
                        "Código": tx.booking_code,
                        "Huésped": tx.guest_name,
                        "Check-in": tx.check_in.strftime("%Y-%m-%d"),
                        "Check-out": tx.check_out.strftime("%Y-%m-%d"),
                        "Noches": tx.nights,
                        "Monto": float(tx.amount),
                        "Moneda": tx.currency,
                        "Estado": tx.status,
                        "Estado Pago": tx.payment_status or "N/A",
                        "Fecha Creación": tx.created_at.strftime("%Y-%m-%d %H:%M"),
                    }
                )

            df = pd.DataFrame(trans_data)

            # Escribir DataFrame en la hoja
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_trans.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                        )
                        cell.alignment = Alignment(horizontal="center")
                    if c_idx == 6:  # Columna de monto
                        cell.number_format = "#,##0.00"

            # Ajustar anchos
            for column in ws_trans.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:  # nosec B110
                        # Ignorar errores al calcular ancho de columna
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_trans.column_dimensions[column_letter].width = adjusted_width

        # Guardar
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
